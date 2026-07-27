"""
pipeline.io
============
Helpers de leitura de arquivos. Lida com idiossincrasias de encoding,
separador, dialeto BR de números e células mescladas no Excel.

Use estas funções em vez de pd.read_csv / pd.read_excel direto.
"""

from __future__ import annotations
from pathlib import Path
from typing import Optional, Union

import pandas as pd

from pipeline.normalize import detect_encoding, detect_sep


def read_csv_smart(
    path: Union[str, Path],
    *,
    decimal_br: bool = False,
    encoding: Optional[str] = None,
    sep: Optional[str] = None,
    **kwargs,
) -> pd.DataFrame:
    """Lê CSV com auto-detecção de encoding e separador."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    if encoding is None:
        encoding = detect_encoding(path)
    if sep is None:
        sep = detect_sep(path, encoding=encoding)

    if decimal_br:
        kwargs.setdefault("decimal", ",")
        kwargs.setdefault("thousands", ".")

    return pd.read_csv(path, encoding=encoding, sep=sep, **kwargs)


def read_excel_safe(path: Union[str, Path], sheet_name=0, **kwargs) -> pd.DataFrame:
    """Wrapper de pd.read_excel com mensagem de erro clara se falhar."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")
    return pd.read_excel(path, sheet_name=sheet_name, **kwargs)


def read_excel_with_merged_fill(
    path: Union[str, Path],
    sheet_name: str,
    header_row: int = 1,
) -> pd.DataFrame:
    """
    Lê uma aba do Excel desfazendo merged cells e propagando o valor da
    célula superior-esquerda para todas as outras células do range.

    Resolve o caso típico de planilhas ANP onde campos como CNPJ, Razão
    Social e Data de Aprovação são mesclados verticalmente entre linhas
    Hidratado/Anidro do mesmo CNPJ. Sem este fix, a leitura padrão do
    pandas vê apenas a primeira linha do range com valor e marca as
    seguintes como NaN, gerando ~38% de falsos missing.

    Parameters
    ----------
    path : str | Path
    sheet_name : str
    header_row : int (0-indexed)
        Linha do header. Default 1 (segunda linha — formato ANP padrão).

    Returns
    -------
    pd.DataFrame com merged cells preenchidas e linhas all-NA removidas.
    """
    import openpyxl

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Aba '{sheet_name}' não encontrada em {path.name}. "
            f"Abas disponíveis: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    # Captura merged ranges ANTES do unmerge (a coleção é mutada durante o loop)
    merged_ranges = list(ws.merged_cells.ranges)

    for mr in merged_ranges:
        top_left = ws.cell(row=mr.min_row, column=mr.min_col).value
        ws.unmerge_cells(str(mr))
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                ws.cell(row=row, column=col).value = top_left

    data = list(ws.iter_rows(values_only=True))
    if header_row >= len(data):
        raise ValueError(
            f"header_row={header_row} excede o número de linhas ({len(data)})"
        )

    header = list(data[header_row])
    body = data[header_row + 1:]
    df = pd.DataFrame(body, columns=header)
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def file_size_mb(path: Union[str, Path]) -> float:
    """Retorna tamanho em MB, ou NaN se arquivo não existir."""
    path = Path(path)
    if not path.exists():
        return float("nan")
    return round(path.stat().st_size / (1024 ** 2), 3)
